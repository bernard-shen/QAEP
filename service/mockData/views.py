#!/usr/bin/python
# -*- coding: UTF-8 -*-
# --Author: Bernard--
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.utils import timezone
from .models import ConnectionInfo, MockHistory
from .business import (
    get_column_list,
    get_secret_list,
    test_connect,
    get_db_list,
    get_table_list,
    MyManualMock
)
from .business.mock_data import fake
import json
import logging
import csv
import io
import uuid
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

logger = logging.getLogger(__name__)

# 移除未使用的变量
resp_correct = {"code": 200, "msg": "请求成功!", "data": []}
resp_wrong = {"code": 404, "msg": "请求失败!", "data": []}

# 基础响应格式
def success_response(data=None, message="请求成功", count=None):
    response = {
        'code': 200,
        'message': message,
        'data': data or []
    }
    if count is not None:
        response['count'] = count
    return JsonResponse(response)

def error_response(message="请求失败"):
    return JsonResponse({
        'code': 400,
        'message': message,
        'data': []
    })


# 获取连接列表
@require_http_methods(["GET"])
def connection_list(request):
    try:
        # 获取查询参数
        connect_id = request.GET.get('connect_id')
        connect_name = request.GET.get('connect_name')
        sql_type = request.GET.get('sql_type')
        host = request.GET.get('host')
        connect_status = request.GET.get('connect_status')
        page = request.GET.get('page', 1)
        page_size = request.GET.get('page_size', 10)

        # 构建查询
        connections = ConnectionInfo.objects.filter(is_deleted=False)
        
        # 应用过滤条件
        if connect_id:
            connections = connections.filter(connect_id=connect_id)
        if connect_name:
            connections = connections.filter(connect_name__icontains=connect_name)
        if sql_type:
            connections = connections.filter(sql_type=sql_type)
        if host:
            connections = connections.filter(host__icontains=host)
        if connect_status:
            connections = connections.filter(connect_status=connect_status)

        # 分页
        paginator = Paginator(connections, page_size)
        connections_page = paginator.page(page)

        # 构建响应数据
        data = [{
            'id': conn.id,
            'connect_id': conn.connect_id,
            'connect_name': conn.connect_name,
            'sql_type': conn.sql_type,
            'host': conn.host,
            'port': conn.port,
            'db': conn.db,
            'username': conn.username,
            'password': conn.password,
            'connect_status': conn.connect_status,
            'create_user': conn.create_user,
            'create_time': conn.create_time.strftime('%Y-%m-%d %H:%M:%S') if conn.create_time else None,
            'update_time': conn.update_time.strftime('%Y-%m-%d %H:%M:%S') if conn.update_time else None,
        } for conn in connections_page]

        return success_response(data=data, count=paginator.count)

    except Exception as e:
        logger.error(f"获取连接列表失败：{str(e)}")
        return error_response(f"获取连接列表失败：{str(e)}")


# 检查连接名是否重复
@require_http_methods(["GET"])
def check_connect_name(request):
    try:
        connect_name = request.GET.get('connect_name')
        if not connect_name:
            return error_response("连接名不能为空")

        exists = ConnectionInfo.objects.filter(
            connect_name=connect_name, 
            is_deleted=False
        ).exists()

        if exists:
            return error_response("连接名已被占用")
        return success_response(message="连接名可用")

    except Exception as e:
        logger.error(f"检查连接名失败：{str(e)}")
        return error_response(f"检查连接名失败：{str(e)}")


# 新建连接
@require_http_methods(["POST"])
def add_connection(request):
    try:
        data = json.loads(request.body)
        
        # 生成连接ID
        data['connect_id'] = str(uuid.uuid4())
        
        # 创建连接记录
        connection = ConnectionInfo.objects.create(
            connect_id=data['connect_id'],
            connect_name=data.get('connect_name'),
            sql_type=data.get('sql_type'),
            host=data.get('host'),
            port=data.get('port'),
            db=data.get('db'),
            username=data.get('username'),
            password=data.get('password'),
            uri=data.get('uri'),
            connect_status='normal',
            create_user=data.get('create_user')
        )

        return success_response(message="连接创建成功")

    except Exception as e:
        logger.error(f"创建连接失败：{str(e)}")
        return error_response(f"创建连接失败：{str(e)}")


# 删除连接
@require_http_methods(["DELETE"])
def delete_connection(request, connect_id):
    try:
        connection = ConnectionInfo.objects.get(connect_id=connect_id, is_deleted=False)
        connection.is_deleted = True
        connection.save()
        
        return success_response(message="连接删除成功")

    except ConnectionInfo.DoesNotExist:
        return error_response("连接不存在")
    except Exception as e:
        logger.error(f"删除连接失败：{str(e)}")
        return error_response(f"删除连接失败：{str(e)}")


# 更新连接
@require_http_methods(["PUT"])
def update_connection(request, connect_id):
    try:
        data = json.loads(request.body)
        connection = ConnectionInfo.objects.get(connect_id=connect_id, is_deleted=False)
        
        # 更新字段
        for field, value in data.items():
            if hasattr(connection, field):
                setattr(connection, field, value)
        
        connection.save()
        return success_response(message="连接更新成功")

    except ConnectionInfo.DoesNotExist:
        return error_response("连接不存在")
    except Exception as e:
        logger.error(f"更新连接失败：{str(e)}")
        return error_response(f"更新连接失败：{str(e)}")


# 测试连接
@require_http_methods(["POST"])
def test_connection(request):
    try:
        data = json.loads(request.body)
        connect_id = data.get('connect_id')
        
        # 获取连接信息
        connection = ConnectionInfo.objects.get(connect_id=connect_id, is_deleted=False)
        
        # 测试连接
        test_result = test_connect(
            sql_type=connection.sql_type,
            connection_info={
                'host': connection.host,
                'port': connection.port,
                'db': connection.db,
                'username': connection.username,
                'password': connection.password
            }
        )
        
        # 更新连接状态
        if test_result == '1':
            connection.connect_status = 'normal'  # 连接成功，状态正常
            connection.save(update_fields=['connect_status'])
            return success_response(message="连接测试成功")
        else:
            connection.connect_status = 'abnormal'  # 连接失败，状态异常
            connection.save(update_fields=['connect_status'])
            return error_response(message="连接测试失败")

    except ConnectionInfo.DoesNotExist:
        return error_response("连接不存在")
    except Exception as e:
        logger.error(f"测试连接失败：{str(e)}")
        # 如果发生异常，也将状态设置为异常
        try:
            if 'connection' in locals():
                connection.connect_status = 'abnormal'
                connection.save(update_fields=['connect_status'])
        except Exception:
            pass
        return error_response(f"测试连接失败：{str(e)}")


# 一键造数
@require_http_methods(["POST"])
def mock_connect_data(request):
    try:
        data = json.loads(request.body)
        if data:
            test_resp = test_connect(sql_type=data['sql_type'], connection_info=data)
            if test_resp == '1':
                # 获取数据之后，根据参数进行判断，是否使用快速造数，还是自定义造数
                new_mock = MyManualMock(connection_info=data)
                resp = new_mock.manual_mock(sql_type=data['sql_type'], data_lines=data['dataLines'])
                if resp == '造数成功':
                    return success_response(message='造数成功')
                else:
                    return error_response(message=resp)
            else:
                return error_response(message='测试连接失败，请确认连接配置是否正确！')
    except Exception as e:
        logger.error(f"造数失败，错误原因：{str(e)}")
        return error_response(message=f"造数失败，错误原因：{str(e)}")


# 获取数据库列表
@require_http_methods(["GET"])
def get_db_list_view(request):
    try:
        # 从 GET 参数中获取数据
        data = {
            'sql_type': request.GET.get('sql_type'),
            'host': request.GET.get('host'),
            'port': request.GET.get('port'),
            'username': request.GET.get('username'),
            'password': request.GET.get('password', ''),  # 可能需要从其他地方获取密码
            'db': request.GET.get('db')
        }
        
        if data['sql_type'] and data['host']:  # 确保必要参数存在
            resp = get_db_list(sql_type=data['sql_type'], connection_info=data)
            if isinstance(resp, list):
                return success_response(data=resp, message='连接成功')
            else:
                return error_response(message='连接失败')
        else:
            return error_response(message='缺少必要参数')
    except Exception as e:
        logger.error(f"获取数据库列表失败：{str(e)}")
        return error_response(message=f"获取数据库列表失败：{str(e)}")


# 获取表列表
@require_http_methods(["GET"])
def get_table_list_view(request):
    try:
        # 从 GET 参数中获取数据
        data = {
            'sql_type': request.GET.get('sql_type'),
            'host': request.GET.get('host'),
            'port': request.GET.get('port'),
            'username': request.GET.get('username'),
            'password': request.GET.get('password', ''),  # 可能需要从其他地方获取密码
            'db': request.GET.get('db')
        }
        
        if data['sql_type'] and data['host'] and data['db']:  # 确保必要参数存在
            resp = get_table_list(
                sql_type=data['sql_type'], 
                connection_info=data, 
                choice_db=data["db"]
            )
            if isinstance(resp, list):
                return success_response(data=resp, message='连接成功')
            else:
                return error_response(message='连接失败')
        else:
            return error_response(message='缺少必要参数')
    except Exception as e:
        logger.error(f"获取表列表失败：{str(e)}")
        return error_response(message=f"获取表列表失败：{str(e)}")


# 获取字段名列表
@require_http_methods(["GET"])
def get_column_list_view(request):
    try:
        column_list = get_column_list()
        list_temp = [{"name": item, "value": item} for item in column_list]
        return success_response(data=list_temp, message='数据获取成功')
    except Exception as e:
        logger.error(f"获取字段名列表失败：{str(e)}")
        return error_response(message=f"获取字段名列表失败：{str(e)}")


# 获取隐私类型列表
@require_http_methods(["GET"])
def get_secret_list_view(request):
    try:
        secret_list = get_secret_list()
        list_temp = [{"name": item, "value": item} for item in secret_list]
        return success_response(data=list_temp, message='数据获取成功')
    except Exception as e:
        logger.error(f"获取隐私类型列表失败：{str(e)}")
        return error_response(message=f"获取隐私类型列表失败：{str(e)}")


@require_http_methods(["POST"])
def download_mock_data(request):
    # 定义临时文件路径
    tmp_csv = 'tmp.csv'
    tmp_xlsx = 'tmp.xlsx'
    
    try:
        # 删除可能存在的旧临时文件
        for tmp_file in [tmp_csv, tmp_xlsx]:
            if os.path.exists(tmp_file):
                os.remove(tmp_file)

        # 验证请求体是否为空
        if not request.body:
            return error_response("请求数据不能为空")
            
        # 尝试解析 JSON
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return error_response("无效的 JSON 格式")
            
        # 验证必要字段
        if 'column_list' not in data:
            return error_response("缺少必要参数：column_list")
            
        column_list = data.get('column_list', [])
        if not isinstance(column_list, list):
            return error_response("column_list 必须是数组")
            
        if not column_list:
            return error_response("column_list 不能为空")
            
        # 获取并验证数据行数
        data_lines = int(data.get('data_lines', 100))
        if data_lines <= 0:
            return error_response("生成行数必须大于0")
            
        # 所有支持的字段列表
        ALLOWED_COLUMNS = [
            # 17种内置隐私类型
            'address',          # 中文地址
            'bank_card',       # 银行卡号
            'email',           # 电子邮件
            'company_name',    # 企业名称
            'name',            # 中文姓名
            'id_no',          # 身份证号
            'id_card',        # 身份证号别名
            'phone_number',    # 手机号
            'phone',          # 手机号别名
            'fix_phone',      # 固定电话
            'post_code',      # 邮政编码
            'car_no',         # 车牌号
            'social_credit_code', # 社会信用代码
            'car_code',       # 汽车车架号
            'passport',       # 护照号
            'tax_code',       # 税务登记证号
            'organization',   # 组织机构代码
            'enterprise_code', # 营业执照代码
            'individual_business', # 单体商户名称
            'officer_card',   # 军官警官证编号

            # 其他常见类型
            'number',         # 随机整数
            'character',      # 随机字符串
            'description',    # 文章
            'change_line_description', # 带换行符的文章
            'create_time',    # 创建时间
            'update_time',    # 更新时间
            'timestamp',      # 时间戳
            'job',           # 职位
            'full_credit_card', # 完整信用卡信息
            'date',          # 年月日
            'weekday',       # 周几
            'time',          # 时分秒
            'timezone',      # 时区
            'country',       # 国家名称
            'province',      # 省份
            'street',        # 街道
            'color',         # 颜色名称
            'file_path',     # 文件路径
            'hostname',      # 主机名
            'url',           # url
            'image_url',     # 图片url
            'ipv4',          # ipv4
            'ipv6',          # ipv6
            'mac_address',   # mac地址
            'user_name',     # 用户名
            'password',      # 密码
            'md5',           # md5
            'profile',       # 档案(完整)
            'simple_profile', # 档案(简单)
            'dictionary',    # 字典
            'list',          # 列表
            'set',           # 集合
            'special_name',  # 少数民族姓名
            'uuid',          # 7位uuid
            'uid'            # uid
        ]
        
        # 验证请求的字段是否都在允许列表中
        invalid_columns = [col for col in column_list if col not in ALLOWED_COLUMNS]
        if invalid_columns:
            return error_response(f"不支持的字段：{', '.join(invalid_columns)}")
            
        file_type = data.get('file_type', 'csv').lower()
        if file_type not in ['csv', 'xlsx']:
            return error_response("不支持的文件类型")
            
        # 生成模拟数据
        mock_data = []
        for i in range(data_lines):
            row_data = []
            for column in column_list:
                try:
                    method_name = f'get_{column.lower()}'
                    method = getattr(fake, method_name)
                    value = method()
                    
                    # 处理返回值
                    if isinstance(value, (dict, list, set)):
                        value = json.dumps(value, ensure_ascii=False)
                    elif value is not None:
                        value = str(value)
                    else:
                        value = ''
                    row_data.append(value)
                    
                except AttributeError:
                    row_data.append('')
                except Exception as e:
                    logger.error(f"生成字段 {column} 数据失败：{str(e)}")
                    row_data.append('')
            mock_data.append(row_data)
        
        # 检查生成的数据
        if not mock_data or not any(any(row) for row in mock_data):
            return error_response("生成的数据为空")

        # 根据文件类型生成响应
        if file_type == 'csv':
            filename = f'mock_data_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
            
            # 写入临时CSV文件
            with open(tmp_csv, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, lineterminator='\n')
                writer.writerow(column_list)
                writer.writerows(mock_data)
            
            try:
                # 检查文件是否存在和可读
                if not os.path.exists(tmp_csv):
                    return error_response("临时文件未生成")
                    
                if not os.access(tmp_csv, os.R_OK):
                    return error_response("临时文件不可读")
                
                file_size = os.path.getsize(tmp_csv)
                if file_size == 0:
                    return error_response("生成的文件为空")
                
                # 创建文件响应
                file_obj = open(tmp_csv, 'rb')
                response = FileResponse(
                    file_obj,
                    as_attachment=True,
                    filename=filename
                )
                response['Content-Type'] = 'text/csv; charset=utf-8-sig'
                response['Content-Length'] = file_size
                response['Access-Control-Expose-Headers'] = 'Content-Disposition, Content-Length'
                
                # 记录造数历史
                MockHistory.objects.create(
                    mock_type='file',
                    sql_type='',
                    is_increment_schema=False,
                    is_increment_table=False,
                    schema_name='',
                    table_name=filename,
                    is_increment_data=True,
                    increment_lines=data_lines
                )
                
                return response
                
            except Exception as e:
                logger.error(f"创建文件响应失败：{str(e)}")
                return error_response(f"创建文件响应失败：{str(e)}")
            
        else:
            filename = f'mock_data_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            # 创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            
            # 写入表头
            for col_num, column in enumerate(column_list, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'] = column
                ws.column_dimensions[col_letter].width = 15
            
            # 写入数据
            for row_num, row_data in enumerate(mock_data, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
            # 保存到临时文件
            wb.save(tmp_xlsx)
            
            try:
                # 检查文件是否存在和可读
                if not os.path.exists(tmp_xlsx):
                    return error_response("临时文件未生成")
                    
                if not os.access(tmp_xlsx, os.R_OK):
                    return error_response("临时文件不可读")
                
                file_size = os.path.getsize(tmp_xlsx)
                if file_size == 0:
                    return error_response("生成的文件为空")
                
                # 创建文件响应
                file_obj = open(tmp_xlsx, 'rb')
                response = FileResponse(
                    file_obj,
                    as_attachment=True,
                    filename=filename
                )
                response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response['Content-Length'] = file_size
                response['Access-Control-Expose-Headers'] = 'Content-Disposition, Content-Length'
                
                # 记录造数历史
                MockHistory.objects.create(
                    mock_type='file',
                    sql_type='',
                    is_increment_schema=False,
                    is_increment_table=False,
                    schema_name='',
                    table_name=filename,
                    is_increment_data=True,
                    increment_lines=data_lines
                )
                
                return response
                
            except Exception as e:
                logger.error(f"创建文件响应失败：{str(e)}")
                return error_response(f"创建文件响应失败：{str(e)}")
            
    except Exception as e:
        logger.error(f"生成文件失败：{str(e)}")
        return error_response(f"生成文件失败：{str(e)}")


@require_http_methods(["POST"])
def preview_mock_data(request):
    try:
        data = json.loads(request.body)
        column_list = data.get('column_list', [])
        data_lines = int(data.get('data_lines', 5))  # 预览默认显示5行
        
        # 生成预览数据
        preview_data = []
        for _ in range(data_lines):
            row_data = {}
            for column in column_list:
                try:
                    method_name = f'get_{column.lower()}'
                    if hasattr(fake, method_name):
                        method = getattr(fake, method_name)
                        value = method()
                        if isinstance(value, (dict, list, set)):
                            value = json.dumps(value, ensure_ascii=False)
                        elif value is not None:
                            value = str(value)
                        else:
                            value = ''
                        row_data[column] = value
                except Exception as e:
                    row_data[column] = ''
            preview_data.append(row_data)
            
        return success_response(data=preview_data)
        
    except Exception as e:
        logger.error(f"生成预览数据失败：{str(e)}")
        return error_response(f"生成预览数据失败：{str(e)}")